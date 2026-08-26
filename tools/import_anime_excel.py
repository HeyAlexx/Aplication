from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_DIR.parents[1] / "Referencias" / "Anime Data V1.xlsx"
ANIME_JSON_PATH = PROJECT_DIR / "data" / "anime.json"
BACKUP_DIR = PROJECT_DIR / "backups" / "manual-imports"

QUARTERS = {
    1: "Winter",
    2: "Spring",
    3: "Summer",
    4: "Fall",
}

FORMATS = {
    "Serie": "Serie",
    "Movie": "Película",
    "OVA": "OVA",
    "Especial": "Especial",
}

PRODUCTION_STATUS = {
    "Emision": "Emisión",
    "Finalizado": "Finalizado",
    "Pausado": "Pausado",
}

VIEWING_STATUS = {
    "Visto": "Visto",
    "No visto": "No Visto",
    "No Visto": "No Visto",
    "Por Ver": "Por Ver",
    "Ignorado": "Ignorado",
}


def clean_text(value: object) -> str:
    return str(value).strip() if value not in (None, "") else ""


def clean_int(value: object, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def build_record(row_number: int, row: tuple[object, ...]) -> dict[str, object]:
    code, name, status, content_type, chapters, seasons, active_season, watching, year, quarter = row[:10]
    code_text = clean_text(code)
    title = clean_text(name)
    format_name = FORMATS.get(clean_text(content_type), "Serie")
    production_status = PRODUCTION_STATUS.get(clean_text(status), "Pendiente")
    viewing_status = VIEWING_STATUS.get(clean_text(watching), "Por Ver")
    emission_year = clean_int(year) or None
    chapters_count = clean_int(chapters)
    seasons_count = max(1, clean_int(seasons, 1))
    emission_season = QUARTERS.get(clean_int(quarter), "")

    return {
        "id": f"anime-excel-{row_number}-{code_text}",
        "sourceCode": code_text,
        "sourceRow": row_number,
        "title": title,
        "type": "anime",
        "categoryGeneral": "Anime",
        "format": format_name,
        "productionStatus": production_status,
        "viewingStatus": viewing_status,
        "genre": "Anime",
        "year": str(emission_year) if emission_year else "",
        "emissionYear": emission_year,
        "emissionSeason": emission_season,
        "chapters": chapters_count,
        "seasonsCount": seasons_count,
        "activeSeason": clean_text(active_season),
        "tags": [format_name, viewing_status, production_status],
        "image": "",
        "description": "Registro importado desde Anime Data V1.xlsx.",
        "featured": False,
        "status": "Activo",
        "source": "excel-anime-data-v1",
    }


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    sheet = workbook["Catalogo"]
    imported: list[dict[str, object]] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] in (None, "") or row[1] in (None, ""):
            continue
        imported.append(build_record(row_number, row))

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = BACKUP_DIR / f"anime-before-excel-{timestamp}.json"
    shutil.copy2(ANIME_JSON_PATH, backup_path)

    result = imported
    ANIME_JSON_PATH.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(json.dumps({
        "preserved": 0,
        "imported": len(imported),
        "total": len(result),
        "backup": str(backup_path),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
